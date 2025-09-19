import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from zipfile import ZipFile
import os

st.title("📊 Processador de Relatórios de Treinamentos")
st.write("Faça o upload do arquivo Excel (.xlsx) para gerar os arquivos CSV.")

# Upload do arquivo
arquivo = st.file_uploader("Selecione o arquivo Excel", type=["xlsx"])

if arquivo:
    linha_cabecalho = 7  # linha onde estão os nomes das colunas

    # Salvar temporariamente no disco
    caminho_temp = os.path.join("uploads", arquivo.name)
    os.makedirs("uploads", exist_ok=True)
    with open(caminho_temp, "wb") as f:
        f.write(arquivo.getbuffer())

    # 1. Ler com Pandas
    df = pd.read_excel(
        caminho_temp,
        header=linha_cabecalho,
        engine="openpyxl"
    )

    # 2. Ler CPFs crus com openpyxl
    wb = load_workbook(caminho_temp, data_only=True)
    ws = wb.active

    coluna_cpf = None
    for idx, cell in enumerate(ws[linha_cabecalho+1], start=1):
        if str(cell.value).strip().upper() == "CPF":
            coluna_cpf = idx
            break

    cpfs = []
    for row in ws.iter_rows(min_row=linha_cabecalho+2, max_row=ws.max_row, min_col=coluna_cpf, max_col=coluna_cpf):
        valor = row[0].value
        if valor is None:
            cpfs.append("")
        else:
            cpfs.append(str(valor).strip())

    df["CPF"] = cpfs[:len(df)]

    # 3. Selecionar colunas necessárias
    colunas_necessarias = ["TREINAMENTO", "CPF", "DATA"]
    df_selecionado = df[colunas_necessarias].copy()
    df_selecionado = df_selecionado.dropna(subset=["CPF"])

    # 4. Filtrar fora "DDS"
    df_selecionado = df_selecionado[~df_selecionado["TREINAMENTO"].str.contains("DDS", case=False, na=False)]

    # 5. Agrupar por tema e data e gerar ZIP
    arquivos_zip = io.BytesIO()
    with ZipFile(arquivos_zip, "w") as zipf:
        for (tema, data), df_tema in df_selecionado.groupby(["TREINAMENTO", "DATA"]):
            df_tema = df_tema.drop_duplicates(subset=["CPF"])

            # Formatar data
            if isinstance(data, pd.Timestamp):
                data_str = data.strftime("%d.%m.%Y")
            else:
                data_str = str(data).replace("/", ".")

            # Nome do arquivo CSV
            nome_arquivo = f"{str(tema).replace('/', '_').replace('\\', '_')}_{data_str}.csv"

            # Conteúdo do CSV
            conteudo = df_tema.rename(columns={"CPF": "userId"})[["userId"]].to_csv(
                index=False, encoding="utf-8", sep=","
            )

            zipf.writestr(nome_arquivo, conteudo)

    arquivos_zip.seek(0)

    # 6. Botão de download
    st.success("✅ Processamento concluído!")
    st.download_button(
        label="⬇️ Baixar arquivos ZIP",
        data=arquivos_zip,
        file_name="arquivos_treinamentos.zip",
        mime="application/zip"
    )

    # 7. Limpar arquivos e memória
    if os.path.exists(caminho_temp):
        os.remove(caminho_temp)  # apaga Excel enviado

    df = None
    df_selecionado = None
    cpfs = None
    arquivos_zip = None

    st.write("⚠️ Os dados foram processados e removidos da memória e do disco.")


